"""
data_loader.py
Load và parse SO_TAY_KTV.xlsx bằng openpyxl.

Cơ chế tự động reload:
  - get_file_mtime(): trả về mtime của file Excel (không cache).
  - load_data(mtime): cache theo mtime — khi file thay đổi, cache tự hết hạn,
                      Streamlit đọc lại file mà không cần restart app.
"""

import os
import openpyxl
import streamlit as st

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_SEARCH_PATHS = (
    os.path.join(_BASE_DIR, "SO_TAY_KTV.xlsx"),
    os.path.join(_BASE_DIR, "tailieu", "SO_TAY_KTV.xlsx"),
)
_REQUIRED_COLS = frozenset({"ten", "buoc"})
_DEFAULT_FOLDER = "Quy trình"


def _find_file() -> str:
    for path in _SEARCH_PATHS:
        if os.path.isfile(path):
            return path
    raise FileNotFoundError(f"Không tìm thấy SO_TAY_KTV.xlsx trong: {_SEARCH_PATHS}")


def get_file_mtime() -> float:
    """Trả về thời điểm chỉnh sửa cuối của file Excel (epoch seconds). Trả về 0.0 nếu không tìm thấy."""
    try:
        return os.path.getmtime(_find_file())
    except FileNotFoundError:
        return 0.0


@st.cache_data(show_spinner="⏳ Đang tải dữ liệu...")
def load_data(_file_mtime: float) -> list[dict]:
    """
    Đọc SO_TAY_KTV.xlsx, trả về list[dict] với keys: 'ten', 'buoc', 'folder'.
    Cache theo _file_mtime — khi file được lưu, mtime thay đổi và cache tự làm mới.
    Raise FileNotFoundError hoặc ValueError nếu có lỗi.
    """
    file_path = _find_file()
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active

    headers = [
        str(c).strip().lower() if (c := cell.value) is not None else ""
        for cell in next(ws.iter_rows(max_row=1))
    ]

    missing = _REQUIRED_COLS - set(headers)
    if missing:
        wb.close()
        raise ValueError(f"Thiếu cột bắt buộc: {', '.join(sorted(missing))}")

    idx = {name: headers.index(name) for name in ("ten", "buoc")}
    idx_folder = headers.index("folder") if "folder" in headers else None

    def _val(row, col_idx: int) -> str:
        v = row[col_idx].value
        return str(v).strip() if v is not None else ""

    rows = []
    for row in ws.iter_rows(min_row=2):
        ten  = _val(row, idx["ten"])
        buoc = _val(row, idx["buoc"])
        if not ten and not buoc:
            continue
        folder = (_val(row, idx_folder) if idx_folder is not None else "") or _DEFAULT_FOLDER
        rows.append({"ten": ten, "buoc": buoc, "folder": folder})

    wb.close()
    return rows
