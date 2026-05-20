"""
error_code_renderer.py
Render Quy_hoach_ma_loi_FPT_Play.xlsx → Accordion Cards tối ưu mobile.
"""

import os
import re
import html as _html

import streamlit as st
import openpyxl

# ── Cấu hình màu sắc badge theo nền tảng ────────────────────────────────────
_SHEET_CONFIG: dict[str, dict] = {
    "SmartTV LG,Sony HTML,SamSung": {
        "icon": "📺", "label": "SmartTV LG,Sony HTML,SamSung",
        "badge_bg": "#EFF6FF", "badge_color": "#005DA3",
        "badge_border": "#BDD7F5", "accent": "#005DA3",
    },
    "Android(all Platform)": {
        "icon": "🤖", "label": "Android(all Platform)",
        "badge_bg": "#EFFAF4", "badge_color": "#1A7A42",
        "badge_border": "#A3DDB8", "accent": "#1A7A42",
    },
    "iOS": {
        "icon": "🍎", "label": "iOS",
        "badge_bg": "#F3F0FF", "badge_color": "#6D28D9",
        "badge_border": "#C4B5FD", "accent": "#6D28D9",
    },
    "Website": {
        "icon": "🌐", "label": "Website",
        "badge_bg": "#FFF5EF", "badge_color": "#C04800",
        "badge_border": "#FFBF99", "accent": "#F26F21",
    },
    "Không hiện mã lỗi_Dịch vụ": {
        "icon": "⚠️", "label": "Không hiện mã lỗi_Dịch vụ",
        "badge_bg": "#FFFBEB", "badge_color": "#92400E",
        "badge_border": "#FCD34D", "accent": "#D97706",
    },
    "Lỗi box 650 hiện hữu": {
        "icon": "📦", "label": "Lỗi box 650 hiện hữu",
        "badge_bg": "#FFF0F0", "badge_color": "#B91C1C",
        "badge_border": "#FCA5A5", "accent": "#DC2626",
    },
}
_DEFAULT_CONFIG = {
    "icon": "🔧", "label": "Khác",
    "badge_bg": "#F0F3F8", "badge_color": "#3A4454",
    "badge_border": "#C8D0DC", "accent": "#8896A5",
}

_EC_CSS = """<style>
.ec-card{background:#fff;border-radius:14px;margin-bottom:10px;overflow:hidden;
  box-shadow:0 2px 10px rgba(0,0,0,0.05);border:1.5px solid #EDF0F5;
  font-family:'Sora',sans-serif;transition:box-shadow 0.2s}
.ec-card:hover{box-shadow:0 4px 16px rgba(0,0,0,0.09)}
.ec-header{display:flex;align-items:center;padding:12px 14px;gap:10px;
  cursor:pointer;user-select:none}
.ec-badge{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;
  border-radius:20px;font-size:0.68rem;font-weight:700;letter-spacing:0.3px;
  white-space:nowrap;flex-shrink:0;border:1px solid transparent}
.ec-code{font-size:1.0rem;font-weight:800;letter-spacing:-0.5px;flex-shrink:0;min-width:54px}
.ec-desc{font-size:0.78rem;color:#5A6070;flex:1;line-height:1.4;overflow:hidden;
  display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical}
.ec-arrow{font-size:0.9rem;color:#B0BBC8;flex-shrink:0;transition:transform 0.25s;margin-left:2px}
.ec-arrow.open{transform:rotate(180deg)}
.ec-body{padding:0 14px 14px;border-top:1px solid #F0F3F8;
  animation:ec-slide-down 0.2s ease-out}
@keyframes ec-slide-down{from{opacity:0;transform:translateY(-6px)}to{opacity:1;transform:translateY(0)}}
.ec-section-title{font-size:0.72rem;font-weight:700;text-transform:uppercase;
  letter-spacing:0.6px;color:#A0AEBB;margin:12px 0 6px}
.ec-bullets,.ec-steps{margin:0 0 4px;padding-left:18px}
.ec-bullets li,.ec-steps li{font-size:0.82rem;color:#3A4454;line-height:1.6;margin-bottom:3px}
li.ec-sub{font-size:0.78rem;color:#7A8899;margin-left:12px;list-style-type:circle}
.ec-para{font-size:0.82rem;color:#3A4454;line-height:1.6;margin:4px 0}
.ec-divider{height:1px;background:#F0F3F8;margin:10px 0}
.ec-empty{text-align:center;color:#B0BBC8;font-size:0.82rem;padding:24px 0;font-family:'Sora',sans-serif}
.ec-count-badge{display:inline-block;background:#F0F3F8;color:#8896A5;font-size:0.72rem;
  font-weight:600;padding:2px 8px;border-radius:10px;margin-left:8px;font-family:'Sora',sans-serif}
@media(max-width:480px){
  .ec-code{font-size:0.9rem}.ec-desc{font-size:0.74rem}
  .ec-header{padding:11px 12px;gap:8px}.ec-body{padding:0 12px 12px}}
</style>"""


# ── Parse dữ liệu ────────────────────────────────────────────────────────────

def _clean(val) -> str:
    return str(val).strip() if val is not None else ""


def _parse_sheet(ws, sheet_name: str) -> list[dict]:
    """Trả về list[dict] với keys: platform, ma_loi, mo_ta, nguyen_nhan, cach_xu_ly."""

    # Sheet đặc biệt: Box 650 (không có cột chuẩn)
    if sheet_name == "Lỗi box 650 hiện hữu":
        rows = []
        for r in range(2, ws.max_row + 1):
            text = _clean(ws.cell(r, 1).value)
            if not text:
                continue
            lines = text.split("\n")
            rows.append({
                "platform": "Box 650", "ma_loi": "",
                "mo_ta": lines[0].strip(),
                "nguyen_nhan": "",
                "cach_xu_ly": "\n".join(lines[1:]).strip() if len(lines) > 1 else "",
            })
        return rows

    # Sheet: Không hiện mã lỗi_Dịch vụ
    if sheet_name == "Không hiện mã lỗi_Dịch vụ":
        return [
            {
                "platform": "Dịch vụ",
                "ma_loi": _clean(ws.cell(r, 1).value),
                "mo_ta":  _clean(ws.cell(r, 2).value),
                "nguyen_nhan": "",
                "cach_xu_ly": _clean(ws.cell(r, 3).value),
            }
            for r in range(2, ws.max_row + 1)
            if _clean(ws.cell(r, 2).value) or _clean(ws.cell(r, 3).value)
        ]

    # Sheet iOS: col 3 = Nguyên nhân, col 4 = Ghi chú thêm, col 5 = Cách xử lý
    if sheet_name == "iOS":
        rows = []
        for r in range(2, ws.max_row + 1):
            ma_loi = _clean(ws.cell(r, 2).value)
            if not ma_loi:
                continue
            nn, ghi_chu, xu_ly = (
                _clean(ws.cell(r, 3).value),
                _clean(ws.cell(r, 4).value),
                _clean(ws.cell(r, 5).value),
            )
            rows.append({
                "platform": _clean(ws.cell(r, 1).value) or "iOS",
                "ma_loi": ma_loi,
                "mo_ta": nn[:80],
                "nguyen_nhan": "\n".join(filter(None, [nn, ghi_chu])),
                "cach_xu_ly": xu_ly,
            })
        return rows

    # Sheet Website: col 4 = chi tiết
    if sheet_name == "Website":
        return [
            {
                "platform": _clean(ws.cell(r, 1).value) or "Website",
                "ma_loi":   _clean(ws.cell(r, 2).value),
                "mo_ta":    _clean(ws.cell(r, 3).value)[:80],
                "nguyen_nhan": _clean(ws.cell(r, 3).value),
                "cach_xu_ly":  _clean(ws.cell(r, 4).value),
            }
            for r in range(2, ws.max_row + 1)
            if _clean(ws.cell(r, 2).value)
        ]

    # SmartTV & Android: cấu trúc chuẩn 4 cột
    rows = []
    for r in range(2, ws.max_row + 1):
        ma_loi = _clean(ws.cell(r, 2).value)
        if not ma_loi:
            continue
        nn = _clean(ws.cell(r, 3).value)
        first_line = nn.split("\n")[0].lstrip("- ").strip() if nn else ""
        rows.append({
            "platform": _clean(ws.cell(r, 1).value) or sheet_name,
            "ma_loi": ma_loi,
            "mo_ta": first_line[:80],
            "nguyen_nhan": nn,
            "cach_xu_ly": _clean(ws.cell(r, 4).value),
        })
    return rows


# ── Format nội dung → HTML ───────────────────────────────────────────────────

_STEP_RE   = re.compile(r"^(B\d+|Bước\s*\d+)\s*[:.]?\s*", re.IGNORECASE)
_BULLET_RE = re.compile(r"^[-•]\s+")
_SUB_RE    = re.compile(r"^\+\s+")


def _format_detail_html(text: str) -> str:
    """Chuyển multi-line text → HTML danh sách có định dạng."""
    if not text:
        return ""

    parts: list[str] = []
    list_type: str | None = None  # "ol" hoặc "ul"

    def _close_list():
        nonlocal list_type
        if list_type:
            parts.append(f"</{list_type}>")
            list_type = None

    def _ensure_list(tag: str):
        nonlocal list_type
        if list_type != tag:
            _close_list()
            parts.append(f"<{tag} class='ec-{'steps' if tag == 'ol' else 'bullets'}'>")
            list_type = tag

    esc = _html.escape

    for raw in text.split("\n"):
        line = raw.strip()
        if not line:
            continue

        if _STEP_RE.match(line):
            _ensure_list("ol")
            parts.append(f"<li>{esc(_STEP_RE.sub('', line))}</li>")
        elif _BULLET_RE.match(line):
            _ensure_list("ul")
            parts.append(f"<li>{esc(_BULLET_RE.sub('', line))}</li>")
        elif _SUB_RE.match(line):
            parts.append(f"<li class='ec-sub'>{esc(_SUB_RE.sub('', line))}</li>")
        else:
            _close_list()
            parts.append(f"<p class='ec-para'>{esc(line)}</p>")

    _close_list()
    return "\n".join(parts)


# ── Render card ──────────────────────────────────────────────────────────────

def _render_card(row: dict, cfg: dict, card_key: str, is_open: bool) -> bool:
    """Render 1 accordion card. Trả về True nếu user vừa click."""
    esc = _html.escape
    code_str = esc(str(row["ma_loi"])) if row["ma_loi"] else "—"

    badge = (
        f'<span class="ec-badge" style="'
        f'background:{cfg["badge_bg"]};color:{cfg["badge_color"]};'
        f'border-color:{cfg["badge_border"]};">'
        f'{cfg["icon"]} {esc(row["platform"])}</span>'
    )
    arrow_cls = "ec-arrow open" if is_open else "ec-arrow"
    header = (
        f'<div class="ec-header">'
        f'  {badge}'
        f'  <span class="ec-code" style="color:{cfg["accent"]};">{code_str}</span>'
        f'  <span class="{arrow_cls}">▾</span>'
        f'</div>'
    )

    body = ""
    if is_open:
        nn_html    = _format_detail_html(row["nguyen_nhan"])
        xu_ly_html = _format_detail_html(row["cach_xu_ly"])
        sections   = []
        if nn_html:
            sections.append(f'<div class="ec-section-title">🔍 Nguyên nhân</div>{nn_html}')
        if nn_html and xu_ly_html:
            sections.append('<div class="ec-divider"></div>')
        if xu_ly_html:
            sections.append(f'<div class="ec-section-title">🛠 Cách xử lý</div>{xu_ly_html}')
        if not sections:
            sections.append('<p class="ec-para" style="color:#B0BBC8;">Chưa có thông tin chi tiết.</p>')
        body = f'<div class="ec-body">{"".join(sections)}</div>'

    border = cfg["accent"] if is_open else "#EDF0F5"
    st.markdown(
        f'<div class="ec-card" style="border-color:{border};">{header}{body}</div>',
        unsafe_allow_html=True,
    )
    return st.button(
        f"{'Thu gọn' if is_open else 'Xem chi tiết'} mã {code_str}",
        key=card_key,
        use_container_width=True,
        type="secondary",
    )


# ── Entry point ──────────────────────────────────────────────────────────────

@st.cache_data(ttl=120)
def _load_error_data(xlsx_path: str) -> dict[str, list[dict]]:
    """Load & parse tất cả sheet. Cache 2 phút."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {name: _parse_sheet(wb[name], name) for name in wb.sheetnames}
    wb.close()
    return result


def render_error_code_accordion(xlsx_path: str, session_prefix: str = "ec") -> None:
    """Render toàn bộ file mã lỗi dưới dạng Accordion Cards tối ưu mobile."""
    if not os.path.isfile(xlsx_path):
        st.warning(f"⚠️ Chưa tìm thấy file: {os.path.basename(xlsx_path)}")
        return

    st.markdown(_EC_CSS, unsafe_allow_html=True)

    all_data    = _load_error_data(xlsx_path)
    sheet_names = list(all_data.keys())

    # ── Session state keys ──
    sk_sheet  = f"{session_prefix}_sheet"
    sk_open   = f"{session_prefix}_open_card"
    sk_search = f"{session_prefix}_search"

    if st.session_state.get(sk_sheet) not in sheet_names:
        st.session_state[sk_sheet] = sheet_names[0]
    st.session_state.setdefault(sk_open, None)

    active_sheet = st.session_state[sk_sheet]

    # ── Tabs chọn sheet ──
    st.markdown(
        "<p style='color:#8896A5;font-size:0.78rem;margin:10px 0 6px;"
        "font-family:Sora,sans-serif;'>📋 Chọn nền tảng / loại thiết bị:</p>",
        unsafe_allow_html=True,
    )
    for i in range(0, len(sheet_names), 3):
        chunk = sheet_names[i:i + 3]
        for col, name in zip(st.columns(len(chunk)), chunk):
            cfg = _SHEET_CONFIG.get(name, _DEFAULT_CONFIG)
            with col:
                if st.button(
                    cfg["label"], key=f"{session_prefix}_tab_{name}",
                    type="primary" if name == active_sheet else "secondary",
                    use_container_width=True,
                ):
                    st.session_state[sk_sheet] = name
                    st.session_state[sk_open]  = None
                    st.rerun()

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── Tìm kiếm ──
    search_kw = st.text_input(
        "", placeholder="🔍  Tìm mã lỗi hoặc từ khóa trong sheet này...",
        label_visibility="collapsed", key=sk_search,
    ).strip().lower()

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── Lọc dữ liệu ──
    rows = all_data[active_sheet]
    if search_kw:
        rows = [
            r for r in rows
            if any(search_kw in str(r[f]).lower() for f in ("ma_loi", "mo_ta", "nguyen_nhan", "cach_xu_ly"))
        ]

    cfg = _SHEET_CONFIG.get(active_sheet, _DEFAULT_CONFIG)
    st.markdown(
        f"<p style='font-weight:700;color:{cfg['accent']};font-size:0.88rem;"
        f"font-family:Sora,sans-serif;margin-bottom:10px;'>"
        f"{active_sheet} <span class='ec-count-badge'>{len(rows)} mục</span></p>",
        unsafe_allow_html=True,
    )

    if not rows:
        st.markdown(
            '<div class="ec-empty">😕 Không tìm thấy kết quả.<br>'
            '<small>Thử từ khóa khác hoặc chọn sheet khác.</small></div>',
            unsafe_allow_html=True,
        )
        return

    # ── Render cards ──
    open_key = st.session_state[sk_open]
    for idx, row in enumerate(rows):
        card_key = f"{session_prefix}_card_{active_sheet}_{idx}"
        is_open  = open_key == card_key
        if _render_card(row, cfg, card_key, is_open):
            st.session_state[sk_open] = None if is_open else card_key
            st.rerun()

    # ── Download ──
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    with open(xlsx_path, "rb") as f:
        st.download_button(
            label="📥  Tải file Excel đầy đủ",
            data=f,
            file_name=os.path.basename(xlsx_path),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{session_prefix}_dl",
        )
