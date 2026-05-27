"""
excel_renderer.py
Render file Excel → HTML table tối ưu mobile, giữ màu sắc từ Excel.
"""

import os
import streamlit as st
import openpyxl

# ── Màu theme mặc định của Office ────────────────────────────────────────────

_OFFICE_THEME: dict[int, tuple[int, int, int]] = {
    0: (0, 0, 0),       1: (255, 255, 255),
    2: (68, 84, 106),   3: (231, 230, 230),
    4: (68, 114, 196),  5: (237, 125, 49),
    6: (165, 165, 165), 7: (255, 192, 0),
    8: (68, 114, 196),  9: (112, 173, 71),
}
_TRANSPARENT = frozenset({"00000000", "FFFFFFFF", "00FFFFFF"})
_MONEY_UNITS = ("K", "đ", "VND", "%")

# ── Style strings dùng chung ──────────────────────────────────────────────────

_S_WRAP = (
    "overflow-x:auto;-webkit-overflow-scrolling:touch;"
    "border-radius:12px;border:1px solid #E4E9F0;"
    "box-shadow:0 2px 12px rgba(0,0,0,0.06);margin-top:8px;"
)
_S_TABLE = "border-collapse:collapse;width:max-content;min-width:100%;font-family:'Sora',sans-serif;"
_S_CELL  = (
    "white-space:nowrap;min-width:72px;padding:9px 14px;"
    "border:1px solid #E4E9F0;vertical-align:middle;"
    "font-size:0.9rem;font-family:'Sora',sans-serif;line-height:1.5;"
)
_S_HEAD  = (
    "white-space:nowrap;min-width:72px;padding:10px 14px;"
    "border:1px solid #D0D8E8;border-bottom:2px solid #B0BEEF;"
    "vertical-align:middle;font-size:0.88rem;font-family:'Sora',sans-serif;"
    "font-weight:700;text-align:center;background-color:#EAF0FB;color:#1a2a3a;"
)
_S_MONEY = "color:#c04800;font-weight:700;"


# ── Helpers màu sắc ───────────────────────────────────────────────────────────

def _hex_to_rgb(hex_str: str) -> tuple[int, int, int] | None:
    if not hex_str or hex_str in _TRANSPARENT:
        return None
    try:
        h = hex_str[-6:]
        return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
    except Exception:
        return None


def _apply_tint(base: tuple[int, int, int], tint: float) -> tuple[int, int, int]:
    r, g, b = base
    if tint > 0:
        return (int(r + (255 - r) * tint), int(g + (255 - g) * tint), int(b + (255 - b) * tint))
    return (int(r * (1 + tint)), int(g * (1 + tint)), int(b * (1 + tint)))


def _resolve_font_color(font) -> tuple[int, int, int] | None:
    if not font or not font.color:
        return None
    fc = font.color
    try:
        if fc.type == "rgb":
            return _hex_to_rgb(fc.rgb)
        if fc.type == "theme":
            base = _OFFICE_THEME.get(int(fc.theme))
            if base is None:
                return None
            tint = getattr(fc, "tint", 0) or 0
            return base if tint == 0 else _apply_tint(base, tint)
    except Exception:
        pass
    return None


def _luminance(rgb: tuple[int, int, int]) -> float:
    return 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]


def _cell_extra_css(cell) -> str:
    """Trả về CSS bổ sung (bg, color, bold, align) từ định dạng ô Excel."""
    parts: list[str] = []
    bg_rgb = None

    try:
        fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
        bg_rgb = _hex_to_rgb(fg) if fg else None
    except Exception:
        pass

    if bg_rgb:
        parts.append(f"background-color:rgb{bg_rgb};")

    try:
        if cell.font and cell.font.bold:
            parts.append("font-weight:700;")
    except Exception:
        pass

    try:
        h = cell.alignment.horizontal if cell.alignment else None
        if h in ("center", "centerContinuous"):
            parts.append("text-align:center;")
        elif h == "right":
            parts.append("text-align:right;")
    except Exception:
        pass

    if bg_rgb:
        parts.append("color:#ffffff;" if _luminance(bg_rgb) < 140 else "color:#1a2a3a;")
    else:
        font_rgb = _resolve_font_color(getattr(cell, "font", None))
        if font_rgb and _luminance(font_rgb) < 230:
            parts.append(f"color:rgb{font_rgb};")

    return "".join(parts)


# ── Render worksheet ──────────────────────────────────────────────────────────

def _render_ws_html(ws) -> str:
    max_row, max_col = ws.max_row, ws.max_column

    # Xây dựng bảng merged cells
    skip: set[tuple[int, int]] = set()
    span: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        r0, r1, c0, c1 = rng.min_row, rng.max_row, rng.min_col, rng.max_col
        span[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
        skip.update(
            (r, c)
            for r in range(r0, r1 + 1)
            for c in range(c0, c1 + 1)
            if (r, c) != (r0, c0)
        )

    # Tìm hàng đầu có dữ liệu
    start_row = next(
        (r for r in range(1, max_row + 1)
         if any(ws.cell(r, c).value is not None for c in range(1, max_col + 1))),
        1,
    )

    # Lọc hàng có nội dung
    content_rows = [
        r for r in range(start_row, max_row + 1)
        if any(
            (r, c) not in skip and (ws.cell(r, c).value is not None or (r, c) in span)
            for c in range(1, max_col + 1)
        )
    ]

    if not content_rows:
        return f'<div style="{_S_WRAP}"><p style="padding:1rem;color:#888">Không có dữ liệu.</p></div>'

    def _cell_tag(r: int, c: int, is_header: bool) -> str:
        cell  = ws.cell(r, c)
        text  = "" if cell.value is None else str(cell.value).strip().replace("\n", "<br>")
        rs, cs = span.get((r, c), (1, 1))
        attrs  = (f' rowspan="{rs}"' if rs > 1 else "") + (f' colspan="{cs}"' if cs > 1 else "")
        extra  = _cell_extra_css(cell)

        if is_header:
            bg_override = extra if "background-color" in extra else ""
            return f'<th{attrs} style="{_S_HEAD}{bg_override}">{text}</th>'

        is_money = (
            text and "background-color" not in extra
            and any(u in text for u in _MONEY_UNITS)
            and any(ch.isdigit() for ch in text)
        )
        money_css = _S_MONEY if is_money else ""
        return f'<td{attrs} style="{_S_CELL}{extra}{money_css}">{text}</td>'

    def _row_html(r: int, is_header: bool) -> str:
        cells = "".join(
            _cell_tag(r, c, is_header)
            for c in range(1, max_col + 1)
            if (r, c) not in skip
        )
        return f"<tr>{cells}</tr>"

    header = _row_html(content_rows[0], is_header=True)
    body   = "".join(_row_html(r, is_header=False) for r in content_rows[1:])

    return (
        f'<div style="{_S_WRAP}">'
        f'<table style="{_S_TABLE}">'
        f"<thead>{header}</thead>"
        f"<tbody>{body}</tbody>"
        f"</table></div>"
    )


# ── Entry point ───────────────────────────────────────────────────────────────

@st.cache_data(ttl=60)
def render_excel_file(file_path: str) -> list[tuple[str, str]]:
    """
    Đọc file Excel, render từng sheet thành HTML.
    Trả về list[(sheet_name, html_string)]. Cache 60 giây.
    """
    if not os.path.isfile(file_path):
        return [("Lỗi", f"<p style='color:red'>❌ Không tìm thấy: <code>{file_path}</code></p>")]
    try:
        wb = openpyxl.load_workbook(file_path)
        result = [(name, _render_ws_html(wb[name])) for name in wb.sheetnames]
        wb.close()
        return result
    except Exception as e:
        return [("Lỗi", f"<p style='color:red'>❌ Lỗi đọc file: {e}</p>")]
